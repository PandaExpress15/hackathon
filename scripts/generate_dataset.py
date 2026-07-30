#!/usr/bin/env python3
"""Generate the deterministic synthetic CareerProof AI demonstration dataset."""

from __future__ import annotations

import argparse
import math
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from careerproof.config import RAW_DATA_PATH, RAW_XLSX_PATH, SYNTHETIC_DISCLOSURE  # noqa: E402

SEED = 20260730
ANCHOR_DATE = pd.Timestamp("2026-07-30")
BASE_ROWS = 640

ROLE_PROFILES: dict[str, dict[str, object]] = {
    "Electrical Engineer": {
        "family": "Electrical & Hardware Engineering",
        "industry": ["Energy", "Advanced Manufacturing", "Aerospace", "Robotics"],
        "skills": ["MATLAB", "AutoCAD", "Circuit Design", "PLC", "C++", "Python", "Altium", "Communication"],
        "preferred": ["Arduino", "Embedded C", "Project Management", "LabVIEW", "Git"],
        "salary": 79000,
        "remote_bias": 0.10,
    },
    "Software Engineer": {
        "family": "Software Engineering",
        "industry": ["Software", "FinTech", "HealthTech", "Cloud Services"],
        "skills": ["Python", "Java", "JavaScript", "Git", "SQL", "Linux", "React", "AWS"],
        "preferred": ["Docker", "Azure", "C++", "Communication", "Project Management"],
        "salary": 98000,
        "remote_bias": 0.55,
    },
    "Data Analyst": {
        "family": "Data & Analytics",
        "industry": ["Consulting", "Healthcare", "Retail", "Financial Services"],
        "skills": ["SQL", "Excel", "Python", "Data Visualization", "Power BI", "Tableau", "Communication"],
        "preferred": ["Statistics", "Project Management", "AWS", "Git"],
        "salary": 72000,
        "remote_bias": 0.45,
    },
    "Cybersecurity Analyst": {
        "family": "Cybersecurity",
        "industry": ["Cybersecurity", "Financial Services", "Government Technology", "Cloud Services"],
        "skills": ["Cybersecurity", "Networking", "Linux", "Python", "SIEM", "Incident Response", "Communication"],
        "preferred": ["AWS", "Azure", "SQL", "Git", "Project Management"],
        "salary": 85000,
        "remote_bias": 0.42,
    },
    "IT Support Specialist": {
        "family": "Information Technology",
        "industry": ["Education", "Healthcare", "Managed Services", "Retail"],
        "skills": ["Technical Support", "Windows", "Networking", "Communication", "Linux", "Ticketing Systems"],
        "preferred": ["Python", "Azure", "Cybersecurity", "Project Management"],
        "salary": 54000,
        "remote_bias": 0.25,
    },
    "Embedded Systems Engineer": {
        "family": "Electrical & Hardware Engineering",
        "industry": ["Robotics", "Automotive", "Aerospace", "Medical Devices"],
        "skills": ["Embedded C", "C++", "Arduino", "Microcontrollers", "Circuit Design", "Git", "Linux"],
        "preferred": ["MATLAB", "Python", "RTOS", "Communication", "Altium"],
        "salary": 91000,
        "remote_bias": 0.16,
    },
    "Automation Engineer": {
        "family": "Electrical & Hardware Engineering",
        "industry": ["Advanced Manufacturing", "Energy", "Logistics", "Food Production"],
        "skills": ["PLC", "Python", "AutoCAD", "Control Systems", "MATLAB", "Communication"],
        "preferred": ["Robotics", "Project Management", "SQL", "Git", "Arduino"],
        "salary": 84000,
        "remote_bias": 0.12,
    },
    "Product Analyst": {
        "family": "Data & Analytics",
        "industry": ["Software", "E-commerce", "FinTech", "Consumer Technology"],
        "skills": ["SQL", "Excel", "Data Visualization", "Communication", "Product Analytics", "Python"],
        "preferred": ["Tableau", "Power BI", "Project Management", "Statistics", "JavaScript"],
        "salary": 78000,
        "remote_bias": 0.48,
    },
    "Network Engineer": {
        "family": "Information Technology",
        "industry": ["Telecommunications", "Managed Services", "Cloud Services", "Government Technology"],
        "skills": ["Networking", "Linux", "Cybersecurity", "Cisco", "Python", "Communication"],
        "preferred": ["AWS", "Azure", "Project Management", "Git", "SIEM"],
        "salary": 80000,
        "remote_bias": 0.24,
    },
    "Junior Developer": {
        "family": "Software Engineering",
        "industry": ["Software", "Digital Agency", "Education Technology", "E-commerce"],
        "skills": ["JavaScript", "Python", "Git", "HTML/CSS", "React", "SQL", "Communication"],
        "preferred": ["Java", "AWS", "Linux", "Project Management", "C++"],
        "salary": 67000,
        "remote_bias": 0.50,
    },
}

ROLE_WEIGHTS = np.array([0.11, 0.16, 0.14, 0.10, 0.08, 0.10, 0.08, 0.08, 0.07, 0.08])
ROLE_WEIGHTS = ROLE_WEIGHTS / ROLE_WEIGHTS.sum()

COMPANIES: dict[str, dict[str, object]] = {
    "Axiom Circuit Labs": {"intern_bias": 1.7, "salary": 1.03, "roles": ["Electrical Engineer", "Embedded Systems Engineer", "Automation Engineer"]},
    "Blue Ridge Analytics": {"intern_bias": 1.3, "salary": 0.98, "roles": ["Data Analyst", "Product Analyst", "Software Engineer"]},
    "Cascade Secure Systems": {"intern_bias": 1.1, "salary": 1.07, "roles": ["Cybersecurity Analyst", "Network Engineer", "Software Engineer"]},
    "Northstar Robotics": {"intern_bias": 1.8, "salary": 1.06, "roles": ["Embedded Systems Engineer", "Automation Engineer", "Software Engineer"]},
    "Harborlight HealthTech": {"intern_bias": 1.2, "salary": 1.01, "roles": ["Software Engineer", "Data Analyst", "IT Support Specialist"]},
    "Keystone Energy Works": {"intern_bias": 1.4, "salary": 1.00, "roles": ["Electrical Engineer", "Automation Engineer", "Data Analyst"]},
    "Rainier Cloud Group": {"intern_bias": 0.9, "salary": 1.10, "roles": ["Software Engineer", "Cybersecurity Analyst", "Network Engineer"]},
    "Hoosier Automation Co.": {"intern_bias": 1.6, "salary": 0.96, "roles": ["Automation Engineer", "Electrical Engineer", "Embedded Systems Engineer"]},
    "Potomac Digital Services": {"intern_bias": 1.1, "salary": 1.00, "roles": ["IT Support Specialist", "Cybersecurity Analyst", "Junior Developer"]},
    "Redwood Product Studio": {"intern_bias": 1.0, "salary": 1.05, "roles": ["Product Analyst", "Software Engineer", "Junior Developer"]},
    "Summit Network Partners": {"intern_bias": 0.8, "salary": 1.02, "roles": ["Network Engineer", "IT Support Specialist", "Cybersecurity Analyst"]},
    "Great Lakes Data Cooperative": {"intern_bias": 1.5, "salary": 0.99, "roles": ["Data Analyst", "Product Analyst", "Software Engineer"]},
    "Sonoran Systems Lab": {"intern_bias": 1.4, "salary": 1.01, "roles": ["Embedded Systems Engineer", "Software Engineer", "Network Engineer"]},
    "Peachtree Technology Group": {"intern_bias": 1.0, "salary": 0.97, "roles": ["IT Support Specialist", "Junior Developer", "Data Analyst"]},
    "Front Range Controls": {"intern_bias": 1.3, "salary": 1.04, "roles": ["Automation Engineer", "Electrical Engineer", "Embedded Systems Engineer"]},
    "Liberty Financial Tech": {"intern_bias": 0.9, "salary": 1.12, "roles": ["Software Engineer", "Cybersecurity Analyst", "Data Analyst"]},
    "Orchard Education Labs": {"intern_bias": 1.7, "salary": 0.93, "roles": ["Junior Developer", "IT Support Specialist", "Data Analyst"]},
    "Pioneer Aerospace Works": {"intern_bias": 1.2, "salary": 1.09, "roles": ["Electrical Engineer", "Embedded Systems Engineer", "Software Engineer"]},
}

LOCATIONS = [
    ("Seattle", "WA", 0.090),
    ("Baltimore", "MD", 0.080),
    ("Indianapolis", "IN", 0.082),
    ("Austin", "TX", 0.078),
    ("Raleigh", "NC", 0.076),
    ("Denver", "CO", 0.070),
    ("Chicago", "IL", 0.068),
    ("Boston", "MA", 0.060),
    ("Atlanta", "GA", 0.065),
    ("Phoenix", "AZ", 0.055),
    ("Portland", "OR", 0.050),
    ("San Diego", "CA", 0.050),
    ("Pittsburgh", "PA", 0.045),
    ("Columbus", "OH", 0.045),
    ("Minneapolis", "MN", 0.043),
    ("Salt Lake City", "UT", 0.043),
]

EXPERIENCE_MULTIPLIERS = {
    "Internship": 0.47,
    "Entry Level": 0.78,
    "Associate": 0.92,
    "Mid Level": 1.12,
    "Senior": 1.38,
}
EXPERIENCE_WEIGHTS = np.array([0.15, 0.32, 0.20, 0.22, 0.11])

FIRST_NAMES = ["Avery", "Jordan", "Taylor", "Morgan", "Riley", "Cameron", "Casey", "Parker", "Quinn", "Skyler", "Reese", "Drew"]
LAST_NAMES = ["Lee", "Patel", "Garcia", "Kim", "Smith", "Nguyen", "Brown", "Wilson", "Davis", "Clark", "Martin", "Lewis"]

TITLE_VARIANTS = {
    "Electrical Engineer": ["Electrical Engineer", "Junior Electrical Engineer", "Power Systems Engineer", "Electrical Design Engineer"],
    "Software Engineer": ["Software Engineer", "Backend Engineer", "Full Stack Engineer", "Application Developer"],
    "Data Analyst": ["Data Analyst", "Business Data Analyst", "Reporting Analyst", "Junior Data Analyst"],
    "Cybersecurity Analyst": ["Cybersecurity Analyst", "SOC Analyst", "Security Operations Analyst", "Information Security Analyst"],
    "IT Support Specialist": ["IT Support Specialist", "Help Desk Technician", "Desktop Support Analyst", "Technical Support Specialist"],
    "Embedded Systems Engineer": ["Embedded Systems Engineer", "Firmware Engineer", "Embedded Software Engineer", "IoT Engineer"],
    "Automation Engineer": ["Automation Engineer", "Controls Engineer", "PLC Engineer", "Industrial Automation Engineer"],
    "Product Analyst": ["Product Analyst", "Product Data Analyst", "Digital Product Analyst", "Associate Product Analyst"],
    "Network Engineer": ["Network Engineer", "Network Operations Engineer", "Infrastructure Engineer", "Junior Network Engineer"],
    "Junior Developer": ["Junior Developer", "Junior Web Developer", "Associate Software Developer", "Entry-Level Developer"],
}

EDUCATION = ["High school diploma", "Associate degree", "Bachelor's degree", "Bachelor's degree preferred", "No degree specified"]


def _pick_work_mode(rng: np.random.Generator, role: str) -> str:
    remote_bias = float(ROLE_PROFILES[role]["remote_bias"])
    remote = min(max(remote_bias, 0.08), 0.62)
    hybrid = 0.34 if remote < 0.40 else 0.29
    onsite = 1.0 - remote - hybrid
    return str(rng.choice(["Remote", "Hybrid", "On-site"], p=[remote, hybrid, onsite]))


def _pick_company(rng: np.random.Generator, role: str) -> str:
    names = list(COMPANIES)
    weights = np.array([3.2 if role in COMPANIES[name]["roles"] else 0.55 for name in names], dtype=float)
    weights /= weights.sum()
    return str(rng.choice(names, p=weights))


def _pick_experience(rng: np.random.Generator, company: str) -> str:
    weights = EXPERIENCE_WEIGHTS.copy()
    weights[0] *= float(COMPANIES[company]["intern_bias"])
    weights /= weights.sum()
    return str(rng.choice(list(EXPERIENCE_MULTIPLIERS), p=weights))


def _skills_for_row(rng: np.random.Generator, role: str, work_mode: str) -> tuple[str, str]:
    profile = ROLE_PROFILES[role]
    core = list(profile["skills"])
    preferred = list(profile["preferred"])

    if work_mode == "Remote":
        for extra in ["AWS", "Azure", "Git", "Communication", "Project Management"]:
            if extra not in core:
                core.append(extra)
    required_count = int(rng.integers(4, min(7, len(core)) + 1))
    preferred_count = int(rng.integers(2, min(5, len(preferred)) + 1))
    required = sorted(rng.choice(core, size=required_count, replace=False).tolist())
    preferred_pick = sorted(rng.choice(preferred, size=preferred_count, replace=False).tolist())
    return " | ".join(required), " | ".join(preferred_pick)


def _salary_for_row(
    rng: np.random.Generator,
    role: str,
    experience: str,
    company: str,
    city: str,
) -> tuple[float | None, float | None]:
    if rng.random() < 0.155:
        return None, None

    base = float(ROLE_PROFILES[role]["salary"])
    multiplier = EXPERIENCE_MULTIPLIERS[experience] * float(COMPANIES[company]["salary"])
    location_multiplier = {
        "Seattle": 1.13,
        "Boston": 1.12,
        "San Diego": 1.10,
        "Austin": 1.06,
        "Denver": 1.05,
        "Baltimore": 1.01,
        "Indianapolis": 0.94,
        "Columbus": 0.93,
        "Pittsburgh": 0.95,
    }.get(city, 1.0)
    midpoint = base * multiplier * location_multiplier * float(rng.normal(1.0, 0.055))
    spread = float(rng.uniform(0.08, 0.18))
    salary_min = round(midpoint * (1.0 - spread) / 1000) * 1000
    salary_max = round(midpoint * (1.0 + spread) / 1000) * 1000
    return max(salary_min, 15000), max(salary_max, salary_min + 1000)


def generate_dataframe(rows: int = BASE_ROWS, seed: int = SEED) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    roles = list(ROLE_PROFILES)
    location_probs = np.array([row[2] for row in LOCATIONS], dtype=float)
    location_probs /= location_probs.sum()

    records: list[dict[str, object]] = []
    for index in range(rows):
        role = str(rng.choice(roles, p=ROLE_WEIGHTS))
        company = _pick_company(rng, role)
        experience = _pick_experience(rng, company)
        work_mode = _pick_work_mode(rng, role)
        city_idx = int(rng.choice(len(LOCATIONS), p=location_probs))
        city, state, _ = LOCATIONS[city_idx]
        date_posted = ANCHOR_DATE - pd.Timedelta(days=int(rng.integers(0, 181)))
        required_skills, preferred_skills = _skills_for_row(rng, role, work_mode)
        salary_min, salary_max = _salary_for_row(rng, role, experience, company, city)
        recruiter_first = str(rng.choice(FIRST_NAMES))
        recruiter_last = str(rng.choice(LAST_NAMES))
        recruiter_name = f"{recruiter_first} {recruiter_last}"
        domain = str(rng.choice(["example.com", "example.org", "example.net"]))
        recruiter_email = f"{recruiter_first.lower()}.{recruiter_last.lower()}{index % 97}@{domain}"
        area_code = {"WA": 206, "MD": 410, "IN": 317, "TX": 512, "NC": 919}.get(state, 555)
        recruiter_phone = f"({area_code}) 555-{100 + (index % 100):04d}"
        source_id = f"SYN-{date_posted:%Y%m%d}-{index + 1:05d}"
        years = {
            "Internship": 0,
            "Entry Level": int(rng.choice([0, 0, 1, 1, 2])),
            "Associate": int(rng.choice([1, 2, 2, 3])),
            "Mid Level": int(rng.choice([3, 4, 5, 6])),
            "Senior": int(rng.choice([6, 7, 8, 10])),
        }[experience]
        title = str(rng.choice(TITLE_VARIANTS[role]))
        industry = str(rng.choice(ROLE_PROFILES[role]["industry"]))
        education = str(rng.choice(EDUCATION, p=[0.07, 0.08, 0.50, 0.22, 0.13]))
        remote_eligible = work_mode in {"Remote", "Hybrid"}
        description = (
            f"Synthetic {experience.lower()} opportunity in {industry.lower()} focused on "
            f"{required_skills.split(' | ')[0]} and collaborative problem solving."
        )
        records.append(
            {
                "posting_id": f"CP-{index + 1:05d}",
                "date_posted": date_posted.date().isoformat(),
                "job_title": title,
                "normalized_role": role,
                "role_family": ROLE_PROFILES[role]["family"],
                "company": company,
                "industry": industry,
                "city": city,
                "state": state,
                "country": "United States",
                "work_mode": work_mode,
                "experience_level": experience,
                "employment_type": "Internship" if experience == "Internship" else str(rng.choice(["Full-time", "Part-time", "Contract"], p=[0.88, 0.05, 0.07])),
                "salary_min": salary_min,
                "salary_max": salary_max,
                "salary_period": "year",
                "salary_currency": "USD",
                "required_skills": required_skills,
                "preferred_skills": preferred_skills,
                "education_requirement": education,
                "years_experience_required": years,
                "remote_eligible": remote_eligible,
                "description_excerpt": description,
                "recruiter_name": recruiter_name,
                "recruiter_email": recruiter_email,
                "recruiter_phone": recruiter_phone,
                "source_type": "synthetic_job_board",
                "source_record_id": source_id,
                "synthetic_record": True,
            }
        )

    frame = pd.DataFrame.from_records(records)

    # Add intentional but controlled quality issues for the Data Quality demo.
    frame.loc[5, "salary_min"] = -5000
    frame.loc[17, "salary_min"] = 120000
    frame.loc[17, "salary_max"] = 90000
    frame.loc[31, "salary_max"] = -1000
    frame.loc[44, "date_posted"] = "not-a-date"
    frame.loc[52, "required_skills"] = ""
    frame.loc[71, "work_mode"] = "Flexible-ish"
    frame.loc[89, "experience_level"] = "New Grad"
    for idx in [12, 29, 63, 111, 164, 233, 312, 407, 508, 599]:
        frame.loc[idx, "recruiter_phone"] = None
    for idx in [22, 118, 207, 318, 422, 521]:
        frame.loc[idx, "education_requirement"] = None

    # Exact duplicates with the same posting ID.
    duplicates = frame.iloc[[3, 47, 101, 188, 301, 415, 522, 603]].copy()
    frame = pd.concat([frame, duplicates], ignore_index=True)

    # Near duplicates with distinct IDs, preserved for visibility rather than removed.
    for offset, source_idx in enumerate([8, 92, 144, 266, 377, 489], start=1):
        near = frame.iloc[source_idx].copy()
        near["posting_id"] = f"CP-ND-{offset:03d}"
        near["source_record_id"] = f"SYN-NEAR-{offset:03d}"
        near["date_posted"] = (pd.to_datetime(near["date_posted"]) + pd.Timedelta(days=1)).date().isoformat()
        frame = pd.concat([frame, pd.DataFrame([near])], ignore_index=True)

    return frame


def write_xlsx(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="Job Postings", index=False)
        dictionary = pd.DataFrame(
            {
                "Field": frame.columns,
                "Description": [
                    "Synthetic unique posting identifier" if c == "posting_id" else
                    "Synthetic recruiter field; masked by the application" if c.startswith("recruiter_") else
                    "Synthetic source identifier" if c == "source_record_id" else
                    "Synthetic demonstration field"
                    for c in frame.columns
                ],
            }
        )
        dictionary.to_excel(writer, sheet_name="Data Dictionary", index=False)
        pd.DataFrame({"Disclosure": [SYNTHETIC_DISCLOSURE]}).to_excel(writer, sheet_name="Read Me", index=False)

    wb = load_workbook(path)
    navy = "102A43"
    green = "14805E"
    light = "E9F5F0"
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in ws.columns:
            values = [str(c.value) if c.value is not None else "" for c in column_cells[:200]]
            width = min(max(max((len(v) for v in values), default=8) + 2, 11), 34)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        if ws.title == "Job Postings":
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName="CareerProofJobPostings", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(table)
        elif ws.title == "Read Me":
            ws["A2"].fill = PatternFill("solid", fgColor=light)
            ws["A2"].font = Font(color=green, bold=True)
            ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[2].height = 60
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=BASE_ROWS)
    parser.add_argument("--seed", type=int, default=SEED)
    parser.add_argument("--csv", type=Path, default=RAW_DATA_PATH)
    parser.add_argument("--xlsx", type=Path, default=RAW_XLSX_PATH)
    args = parser.parse_args()

    frame = generate_dataframe(rows=args.rows, seed=args.seed)
    args.csv.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(args.csv, index=False)
    write_xlsx(frame, args.xlsx)
    salary_coverage = 1.0 - frame[["salary_min", "salary_max"]].isna().all(axis=1).mean()
    print(f"Wrote {len(frame):,} synthetic rows to {args.csv}")
    print(f"Wrote formatted workbook to {args.xlsx}")
    print(f"Salary coverage: {salary_coverage:.1%}")
    print(SYNTHETIC_DISCLOSURE)


if __name__ == "__main__":
    main()
